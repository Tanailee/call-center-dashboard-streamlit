from __future__ import annotations

import csv
import json
import re
import shutil
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


APP_DIR = Path(__file__).resolve().parents[1]
DOCUMENTS_DIR = APP_DIR.parent
DATA_DIR = APP_DIR / "data"
UPLOADS_DIR = APP_DIR / "uploads"
DEFAULT_DATASET_PATH = UPLOADS_DIR / "current_general_dataset.xlsx"
DEFAULT_COMPLAINT_DATASET_PATH = UPLOADS_DIR / "current_complaint_dataset.xlsx"
OUTPUT_PATH = DATA_DIR / "call-center-dashboard.json"
ACTIVE_DATASET_STATE_PATH = DATA_DIR / "active_dataset.json"
ACTIVE_COMPLAINT_DATASET_STATE_PATH = DATA_DIR / "active_complaint_dataset.json"
SUPPORTED_EXTENSIONS = {".xlsx", ".csv"}
UNKNOWN_PERIOD_KEY = "unknown"

GENERAL_REQUIRED_COLUMN_GROUPS: list[tuple[str, tuple[str, ...]]] = [
    ("Created date/time", ("Created", "Call Date")),
    ("Agent identifier", ("Inputted By",)),
    ("Service", ("Service",)),
    ("Receiving channel", ("Receiving Channel", "Channel")),
    ("Call type", ("Call Type", "Type (In/Out)")),
]
COMPLAINT_REQUIRED_COLUMN_GROUPS: list[tuple[str, tuple[str, ...]]] = [
    ("Complaint category", ("Category", "Complaint Category", "Issue", "Sub Category", "Breakdown")),
    ("Complaint channel", ("Receiving Channels", "Receiving Channel", "Channel")),
]
COMPLAINT_CREATED_CANDIDATES = ("Created", "Created Date", "Updated")
COMPLAINT_MONTH_CANDIDATES = ("Month",)
COMPLAINT_CATEGORY_CANDIDATES = ("Category", "Complaint Category", "Issue", "Sub Category", "Breakdown")
COMPLAINT_CATEGORY_SOURCE_CANDIDATES = ("Category", "Complaint Category")
COMPLAINT_SUBCATEGORY_CANDIDATES = ("Sub Category", "SubCategory", "Sub-Category")
COMPLAINT_CHANNEL_CANDIDATES = ("Receiving Channels", "Receiving Channel", "Channel")
COMPLAINT_STATUS_CANDIDATES = ("Status", "Complaint Status")
COMPLAINT_DETAIL_CANDIDATES = ("Breakdown", "Detail", "Sub Category")
COMPLAINT_SIDE_ISSUE_CANDIDATES = ("Side issue", "Side Issue", "Issue Side")
COMPLAINT_LEVEL_CANDIDATES = ("Complaint Level", "Level", "Severity")
COMPLAINT_OWNER_CANDIDATES = ("Business Owner", "Business owner", "Owner")
COMPLAINT_GENDER_CANDIDATES = ("Gender",)
COMPLAINT_FEEDBACK_CANDIDATES = ("Complaint/Feedback", "Complaint Feedback")
COMPLAINT_RESOLUTION_CANDIDATES = ("Resolution",)
COMPLAINT_SERVICE_CANDIDATES = ("Service",)
COMPLAINT_SIGNAL_CANDIDATES = (
    "Category",
    "Sub Category",
    "Issue",
    "Breakdown",
    "Receiving Channels",
    "Status",
    "Complaint Level",
    "Side issue",
    "Business Owner",
    "Complaint/Feedback",
)
COMPLAINT_FIXED_TAXONOMY: list[tuple[str, list[str]]] = [
    (
        "Mobile App",
        [
            "OTP Failure",
            "App Maintenance/Connection",
            "App Function Issued",
            "App Activate Issue",
            "E-Commerce (X-Pay)",
            "App Login Issue",
            "Message Error",
            "Remittance",
            "Bakong",
            "Phone Topup",
            "KHQR",
            "Bill Payment",
            "Dup-Trx",
            "Feedback on App interface",
        ],
    ),
    (
        "Internet Banking",
        [
            "IB/CIB User Maintenance",
            "Token Error",
            "CIB Function issue",
            "IB Function Issue",
        ],
    ),
    (
        "Card",
        [
            "Card Capture",
            "Cash Dispute",
            "Card Issue",
            "POS Dispute",
            "POS Machine issue",
            "Online Payment Dispute",
            "ATM/CRM/VTM/SEM Machine",
        ],
    ),
    (
        "Marketing",
        [
            "Issued with Website",
            "Promotion/Campaign",
        ],
    ),
    (
        "Process & Policy",
        [
            "Fee & Charge",
            "Term & Condition",
            "Document & Requirement",
            "Loan Process",
            "ETC Usage",
        ],
    ),
    (
        "Customer Services",
        [
            "Staffs Behavior",
            "Customer Relationship",
            "Turn Around Time",
            "Knowledge/Human Error",
            "Branch Look & Feel",
        ],
    ),
]
COMPLAINT_DEFAULT_DETAIL_BY_ISSUE = {
    "Mobile App": "App Maintenance/Connection",
    "Internet Banking": "IB Function Issue",
    "Card": "Card Issue",
    "Marketing": "Promotion/Campaign",
    "Process & Policy": "Term & Condition",
    "Customer Services": "Customer Relationship",
}
COMPLAINT_ISSUE_ALIAS_PATTERNS = {
    "Mobile App": ["mobile app", "mobile banking", "technical issue mb", "technical issue mobile banking", "mb"],
    "Internet Banking": ["internet banking", "ib", "cib"],
    "Card": ["bank card issued", "card", "debit card", "atm card"],
    "Marketing": ["marketing", "promotion", "campaign", "website"],
    "Process & Policy": ["process policy", "process & policy", "policy", "process"],
    "Customer Services": ["customer services", "customer service", "customer issued"],
}
COMPLAINT_DETAIL_ALIAS_PATTERNS = {
    "Mobile App": {
        "OTP Failure": ["otp failure", "otp issue", "otp"],
        "App Maintenance/Connection": ["app maintenance", "maintenance", "app connection", "connection issue", "connection"],
        "App Function Issued": ["app function issued", "app function issue", "function issue"],
        "App Activate Issue": ["app activate issue", "app activation", "activation issue", "activate issue"],
        "E-Commerce (X-Pay)": ["e commerce x pay", "e commerce", "x pay", "xpay"],
        "App Login Issue": ["app login issue", "mobile login issue"],
        "Message Error": ["message error", "error message"],
        "Remittance": ["remittance", "itt", "ott", "cross border"],
        "Bakong": ["bakong"],
        "Phone Topup": ["phone topup fail", "phone topup", "top up", "topup"],
        "KHQR": ["khqr", "kh qr"],
        "Bill Payment": ["bill payment", "biller"],
        "Dup-Trx": ["dup trx", "duplicate trx", "duplicate transaction"],
        "Feedback on App interface": ["app interface", "interface feedback", "ui", "ux"],
    },
    "Internet Banking": {
        "IB/CIB User Maintenance": ["user maintenance", "user profile maintenance", "ib cib user maintenance"],
        "Token Error": ["token error", "soft token", "hard token", "token"],
        "CIB Function issue": ["cib function issue", "cib functions issue"],
        "IB Function Issue": ["ib function issue", "ib functions issue"],
    },
    "Card": {
        "Card Capture": ["card capture"],
        "Cash Dispute": ["cash dispute"],
        "Card Issue": ["card issue", "card issued", "card problem"],
        "POS Dispute": ["pos dispute"],
        "POS Machine issue": ["pos machine issue", "pos machine"],
        "Online Payment Dispute": ["online payment dispute", "online payment"],
        "ATM/CRM/VTM/SEM Machine": ["atm crm vtm sem machine", "atm", "crm", "vtm", "sem", "machine"],
    },
    "Marketing": {
        "Issued with Website": ["issue with website", "issued with website", "website"],
        "Promotion/Campaign": ["promotion", "campaign", "marketing"],
    },
    "Process & Policy": {
        "Fee & Charge": ["fee charge", "fee & charge", "charge"],
        "Term & Condition": ["term condition", "terms condition", "terms", "condition"],
        "Document & Requirement": ["document requirement", "document & requirement", "requirement", "document"],
        "Loan Process": ["loan process", "loan"],
        "ETC Usage": ["etc usage", "etc"],
    },
    "Customer Services": {
        "Staffs Behavior": ["staffs behavior", "staff behaviour", "staff behavior"],
        "Customer Relationship": ["customer relationship", "customer issued", "customer issue", "relationship"],
        "Turn Around Time": ["turn around time", "turnaround time", "tat"],
        "Knowledge/Human Error": ["knowledge human error", "human error", "knowledge"],
        "Branch Look & Feel": ["branch look feel", "look feel", "branch experience"],
    },
}
INQUIRY_CATEGORY_CANDIDATES = (
    "INQUIRY CATEGORY",
    "Inquiry Category",
    "Inquiry category",
    "Category",
)


class DataValidationError(Exception):
    def __init__(self, public_message: str, details: list[str] | None = None) -> None:
        super().__init__(public_message)
        self.public_message = public_message
        self.details = details or []


@dataclass(frozen=True)
class DatasetSource:
    path: Path
    label: str


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split())


def normalize_lookup_text(value: Any) -> str:
    text = clean_text(value).casefold()
    if not text:
        return ""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", text)).strip()


COMPLAINT_TAXONOMY_MAP = {issue: list(details) for issue, details in COMPLAINT_FIXED_TAXONOMY}
COMPLAINT_ALL_ISSUES = [issue for issue, _details in COMPLAINT_FIXED_TAXONOMY]
COMPLAINT_ALL_DETAILS = [detail for _issue, details in COMPLAINT_FIXED_TAXONOMY for detail in details]
COMPLAINT_DETAIL_TO_ISSUE = {
    detail: issue
    for issue, details in COMPLAINT_FIXED_TAXONOMY
    for detail in details
}
COMPLAINT_NORMALIZED_ISSUE_ALIASES = {
    issue: [normalize_lookup_text(alias) for alias in aliases if normalize_lookup_text(alias)]
    for issue, aliases in COMPLAINT_ISSUE_ALIAS_PATTERNS.items()
}
COMPLAINT_NORMALIZED_DETAIL_ALIASES = {
    issue: {
        detail: [normalize_lookup_text(alias) for alias in aliases if normalize_lookup_text(alias)]
        for detail, aliases in detail_aliases.items()
    }
    for issue, detail_aliases in COMPLAINT_DETAIL_ALIAS_PATTERNS.items()
}


def as_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d-%b-%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%m/%d/%Y %H:%M",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def as_number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def get_value(row: dict[str, Any], *candidates: str) -> Any:
    for candidate in candidates:
        if candidate in row and row[candidate] not in (None, ""):
            return row[candidate]
    return None


def normalize_category_value(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if text.casefold() in {"nan", "null", "none"}:
        return ""
    return text


def normalize_call_type(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return "unknown"
    lowered = text.casefold()
    if lowered in {"nan", "null", "none"}:
        return "unknown"
    if lowered.startswith("in"):
        return "inbound"
    if lowered.startswith("out"):
        return "outbound"
    return "unknown"


def normalize_receiving_channel(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return "Unknown Channel"
    if text.casefold() in {"nan", "null", "none"}:
        return "Unknown Channel"
    return text


def normalize_agent_label(value: Any) -> str:
    if value is None:
        return "Unknown Agent"

    if isinstance(value, int):
        cleaned = str(value)
    elif isinstance(value, float):
        cleaned = str(int(value)) if value.is_integer() else format(value, "f").rstrip("0").rstrip(".")
    else:
        cleaned = clean_text(value)

    if not cleaned:
        return "Unknown Agent"
    if cleaned.casefold() in {"nan", "null", "none"}:
        return "Unknown Agent"

    numeric_match = re.fullmatch(r"([+-]?\d+)\.0+", cleaned)
    if numeric_match:
        cleaned = numeric_match.group(1)

    return f"Agent{cleaned}"


def normalize_dimension_label(value: Any, *, fallback: str = "Unknown") -> str:
    text = clean_text(value)
    if not text:
        return fallback
    if text.casefold() in {"nan", "null", "none"}:
        return fallback
    return text


def has_dimension_value(value: Any) -> bool:
    return normalize_dimension_label(value, fallback="") != ""


def normalize_side_issue_label(value: Any) -> str:
    text = normalize_dimension_label(value, fallback="Unknown")
    lowered = text.casefold()
    if lowered == "cnb side":
        return "CNB Side"
    if lowered == "nbc side":
        return "NBC Side"
    if lowered == "customer side":
        return "Customer Side"
    if lowered in {"other bank side", "other side"}:
        return "Other Bank Side"
    if lowered in {"3rd processor", "3rd processors", "3rd procecesors"}:
        return "3rd Processor"
    return text


def normalize_status_label(value: Any) -> str:
    text = normalize_dimension_label(value, fallback="Unknown")
    lowered = text.casefold()
    if lowered in {"close", "closed"}:
        return "Close"
    if lowered in {"progress", "in progress"}:
        return "Progress"
    return text


def normalize_level_label(value: Any) -> str:
    text = normalize_dimension_label(value, fallback="Unknown")
    lowered = text.casefold()
    if lowered == "critical":
        return "Critical"
    if lowered == "high":
        return "High"
    if lowered == "medium":
        return "Medium"
    if lowered == "low":
        return "Low"
    return text


def normalize_gender_label(value: Any) -> str:
    text = normalize_dimension_label(value, fallback="Unknown")
    lowered = text.casefold()
    if lowered in {"m", "male"}:
        return "Male"
    if lowered in {"f", "female"}:
        return "Female"
    return text


def normalize_complaint_feedback_label(value: Any) -> str:
    text = normalize_dimension_label(value, fallback="Unknown")
    lowered = text.casefold()
    if lowered.startswith("complaint"):
        return "Complaint"
    if lowered.startswith("feedback"):
        return "Feedback"
    return text


def normalize_matrix_side_issue_label(value: Any) -> str:
    label = normalize_side_issue_label(value)
    lowered = label.casefold()
    if lowered == "cnb side":
        return "CNB"
    if lowered == "nbc side":
        return "NBC"
    if lowered == "customer side":
        return "Customer Side"
    if lowered == "other bank side":
        return "Other Bank"
    if lowered == "3rd processor":
        return "3rd Processors"
    return "Unknown"


def detect_complaint_issue_label(source_values: list[str], combined_text: str) -> str | None:
    non_empty_values = [value for value in source_values if value]
    for value in non_empty_values:
        for issue, aliases in COMPLAINT_NORMALIZED_ISSUE_ALIASES.items():
            if value in aliases:
                return issue
    for issue, aliases in COMPLAINT_NORMALIZED_ISSUE_ALIASES.items():
        if any(alias and alias in combined_text for alias in aliases):
            return issue
    return None


def detect_complaint_detail_label(
    source_values: list[str],
    combined_text: str,
    preferred_issue: str | None = None,
) -> tuple[str | None, str | None]:
    non_empty_values = [value for value in source_values if value]
    candidate_issues = [preferred_issue] if preferred_issue else []
    candidate_issues.extend(issue for issue in COMPLAINT_ALL_ISSUES if issue != preferred_issue)

    for value in non_empty_values:
        for issue in candidate_issues:
            detail_aliases = COMPLAINT_NORMALIZED_DETAIL_ALIASES.get(issue, {})
            for detail in COMPLAINT_TAXONOMY_MAP.get(issue, []):
                aliases = detail_aliases.get(detail, [])
                if value in aliases:
                    return issue, detail

    for issue in candidate_issues:
        detail_aliases = COMPLAINT_NORMALIZED_DETAIL_ALIASES.get(issue, {})
        for detail in COMPLAINT_TAXONOMY_MAP.get(issue, []):
            aliases = detail_aliases.get(detail, [])
            if any(alias and alias in combined_text for alias in aliases):
                return issue, detail

    return None, None


def map_complaint_record_to_taxonomy(
    *,
    issue_value: Any,
    detail_value: Any,
    category_value: Any,
    subcategory_value: Any,
    service_value: Any,
) -> tuple[str, str, bool]:
    lookup_values = [
        normalize_lookup_text(detail_value),
        normalize_lookup_text(subcategory_value),
        normalize_lookup_text(category_value),
        normalize_lookup_text(issue_value),
        normalize_lookup_text(service_value),
    ]
    combined_text = " | ".join(value for value in lookup_values if value)

    exact_issue = detect_complaint_issue_label(lookup_values[2:], combined_text)
    exact_detail_issue, exact_detail = detect_complaint_detail_label(lookup_values[:4], combined_text, exact_issue)
    if exact_detail_issue and exact_detail:
        return exact_detail_issue, exact_detail, False

    if exact_issue:
        return exact_issue, COMPLAINT_DEFAULT_DETAIL_BY_ISSUE[exact_issue], True

    fallback_detail_issue, fallback_detail = detect_complaint_detail_label(lookup_values, combined_text)
    if fallback_detail_issue and fallback_detail:
        return fallback_detail_issue, fallback_detail, False

    return "Customer Services", COMPLAINT_DEFAULT_DETAIL_BY_ISSUE["Customer Services"], True


def as_month_key(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")

    text = clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d{4}-\d{2}", text):
        return text
    if re.fullmatch(r"\d{4}/\d{2}", text):
        year, month = text.split("/")
        return f"{year}-{month}"

    for fmt in ("%b %Y", "%B %Y", "%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.strftime("%Y-%m")
        except ValueError:
            continue
    return None


def build_display_channel_mix(channel_counter: Counter[str]) -> list[dict[str, Any]]:
    return [{"label": label, "value": value} for label, value in channel_counter.most_common()]


def period_meta_from_key(period_key: str) -> dict[str, Any]:
    if period_key == UNKNOWN_PERIOD_KEY:
        return {
            "key": UNKNOWN_PERIOD_KEY,
            "label": "Unknown",
            "shortLabel": "Unknown",
            "isUnknown": True,
        }

    period_dt = datetime.strptime(period_key, "%Y-%m")
    return {
        "key": period_key,
        "label": period_dt.strftime("%b %Y"),
        "shortLabel": period_dt.strftime("%b"),
        "isUnknown": False,
        "datetime": period_dt,
    }


def supported_dataset_text() -> str:
    return ", ".join(sorted(SUPPORTED_EXTENSIONS))


def required_columns_text(mode: str = "general") -> str:
    if mode == "complaint":
        return (
            "At least one complaint field such as Issue, Breakdown, Status, Complaint Level, "
            "Receiving Channels, Business Owner, Side issue, or Complaint/Feedback."
        )
    groups = GENERAL_REQUIRED_COLUMN_GROUPS
    return "; ".join(f"{label}: {' or '.join(options)}" for label, options in groups)


def invalid_format_error(extra_details: list[str] | None = None, *, mode: str = "general") -> DataValidationError:
    details = list(extra_details or [])
    details.append(f"Required columns: {required_columns_text(mode)}")
    return DataValidationError(
        "Uploaded file format is invalid. Please upload a file matching the required dataset structure.",
        details,
    )


def matching_header(headers: list[str], candidates: tuple[str, ...]) -> str | None:
    normalized = {clean_text(header).casefold(): header for header in headers}
    for candidate in candidates:
        match = normalized.get(clean_text(candidate).casefold())
        if match:
            return match
    return None


def load_xlsx_rows(dataset_path: Path, *, mode: str = "general") -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(dataset_path, data_only=True)
    if mode == "complaint" and "Complaints" in workbook.sheetnames:
        sheet = workbook["Complaints"]
    else:
        sheet = workbook["Clean_Data"] if "Clean_Data" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    headers = [clean_text(cell.value) or f"Column_{index}" for index, cell in enumerate(sheet[1], start=1)]

    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def load_csv_rows(dataset_path: Path) -> list[dict[str, Any]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            with dataset_path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                try:
                    raw_headers = next(reader)
                except StopIteration as error:
                    raise invalid_format_error(["The dataset is empty."]) from error

                headers = [clean_text(header) or f"Column_{index}" for index, header in enumerate(raw_headers, start=1)]
                rows: list[dict[str, Any]] = []

                for values in reader:
                    if not any(clean_text(value) for value in values):
                        continue
                    padded_values = list(values) + [""] * max(0, len(headers) - len(values))
                    rows.append(dict(zip(headers, padded_values[: len(headers)])))
                return rows
        except UnicodeDecodeError as error:
            last_error = error
            continue
    raise invalid_format_error(["The CSV file could not be read. Please save it as UTF-8 or Excel CSV."]) from last_error


def load_rows(dataset_path: Path, *, mode: str = "general") -> list[dict[str, Any]]:
    suffix = dataset_path.suffix.lower()
    if suffix == ".xlsx":
        return load_xlsx_rows(dataset_path, mode=mode)
    if suffix == ".csv":
        return load_csv_rows(dataset_path)
    raise invalid_format_error([f"Supported file types: {supported_dataset_text()}."])


def validate_rows(rows: list[dict[str, Any]], *, mode: str = "general") -> None:
    if not rows:
        raise invalid_format_error(["The dataset is empty."], mode=mode)

    headers = list(rows[0].keys())
    if mode == "general":
        required_groups = GENERAL_REQUIRED_COLUMN_GROUPS
        missing_groups = [label for label, options in required_groups if not matching_header(headers, options)]
        if missing_groups:
            detail = "Missing required columns for: " + ", ".join(missing_groups) + "."
            raise invalid_format_error([detail], mode=mode)

        created_header = matching_header(headers, ("Created", "Call Date"))
        valid_created_values = sum(1 for row in rows if as_datetime(row.get(created_header or "")) is not None)
        if valid_created_values == 0:
            raise invalid_format_error(
                ["The dataset must contain at least one valid date/time value in the Created or Call Date column."],
                mode=mode,
            )
        return

    if not any(matching_header(headers, (candidate,)) for candidate in COMPLAINT_SIGNAL_CANDIDATES):
        raise invalid_format_error(
            [
                "The complaint dataset must include at least one complaint field such as Issue, Breakdown, Status, Complaint Level, Receiving Channels, Business Owner, Side issue, or Complaint/Feedback."
            ],
            mode=mode,
        )


def load_and_validate_rows(dataset_path: Path, *, mode: str = "general") -> list[dict[str, Any]]:
    if not dataset_path.exists():
        raise invalid_format_error(["The selected file could not be found."], mode=mode)
    if dataset_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise invalid_format_error([f"Supported file types: {supported_dataset_text()}."], mode=mode)
    if dataset_path.stat().st_size == 0:
        raise invalid_format_error(["The uploaded file is empty."], mode=mode)

    rows = load_rows(dataset_path, mode=mode)
    validate_rows(rows, mode=mode)
    return rows


def build_payload(rows: list[dict[str, Any]], source_label: str) -> dict[str, Any]:
    headers = list(rows[0].keys()) if rows else []
    category_header = matching_header(headers, INQUIRY_CATEGORY_CANDIDATES)
    receiving_channel_header = matching_header(headers, ("Receiving Channel",))
    created_values = [
        as_datetime(get_value(row, "Created", "Call Date"))
        for row in rows
        if as_datetime(get_value(row, "Created", "Call Date")) is not None
    ]
    unique_dates = sorted({value.date() for value in created_values})

    total_calls = len(rows)
    answered_calls = 0
    missed_calls = 0
    handling_minutes: list[float] = []
    hourly_counter: Counter[str] = Counter()
    daily_counter: Counter[str] = Counter()
    monthly_counter: Counter[str] = Counter()
    service_counter: Counter[str] = Counter()
    channel_counter: Counter[str] = Counter()
    trend_channel_counter: Counter[str] = Counter()
    call_type_counter: Counter[str] = Counter()
    category_counter: Counter[str] = Counter()
    uncategorized_inquiries = 0
    monthly_direction: defaultdict[str, dict[str, int]] = defaultdict(
        lambda: {"inbound": 0, "outbound": 0, "unknown": 0}
    )
    direction_by_month_service: defaultdict[tuple[str, str], dict[str, int]] = defaultdict(
        lambda: {"inbound": 0, "outbound": 0, "unknown": 0}
    )
    direction_totals_by_service: defaultdict[str, dict[str, int]] = defaultdict(
        lambda: {"inbound": 0, "outbound": 0, "unknown": 0, "total": 0}
    )
    agent_service_counter: defaultdict[str, Counter[str]] = defaultdict(Counter)
    agent_totals: defaultdict[str, dict[str, Any]] = defaultdict(
        lambda: {"calls": 0, "answered": 0, "inbound": 0, "outbound": 0, "unknown": 0, "handling": []}
    )

    for row in rows:
        created = as_datetime(get_value(row, "Created", "Call Date"))
        agent = normalize_agent_label(get_value(row, "Inputted By"))
        service = clean_text(get_value(row, "Service")) or "Unspecified"
        channel = clean_text(get_value(row, "Receiving Channel", "Channel")) or "Unspecified"
        call_type = normalize_call_type(get_value(row, "Call Type", "Type (In/Out)"))
        category = normalize_category_value(get_value(row, category_header)) if category_header else ""
        answered_flag = as_number(get_value(row, "Answered Flag"))
        missed_flag = as_number(get_value(row, "Missed Flag"))
        handling = as_number(get_value(row, "Handling Minutes", "Resolution Minutes", "Response Minutes"))
        trend_channel = (
            normalize_receiving_channel(row.get(receiving_channel_header)) if receiving_channel_header else "Unknown Channel"
        )

        answered_calls += int(answered_flag > 0)
        missed_calls += int(missed_flag > 0)
        handling_minutes.append(handling)
        service_counter[service] += 1
        channel_counter[channel] += 1
        trend_channel_counter[trend_channel] += 1
        call_type_counter[call_type] += 1
        direction_totals_by_service[service][call_type] += 1
        direction_totals_by_service[service]["total"] += 1
        agent_service_counter[agent][service] += 1
        if category:
            category_counter[category] += 1
        else:
            uncategorized_inquiries += 1

        summary = agent_totals[agent]
        summary["calls"] += 1
        summary["answered"] += int(answered_flag > 0)
        summary["handling"].append(handling)
        if call_type == "inbound":
            summary["inbound"] += 1
        elif call_type == "outbound":
            summary["outbound"] += 1
        else:
            summary["unknown"] += 1

        month_key = created.strftime("%Y-%m") if created else UNKNOWN_PERIOD_KEY
        monthly_counter[month_key] += 1
        if call_type == "inbound":
            monthly_direction[month_key]["inbound"] += 1
            direction_by_month_service[(month_key, service)]["inbound"] += 1
        elif call_type == "outbound":
            monthly_direction[month_key]["outbound"] += 1
            direction_by_month_service[(month_key, service)]["outbound"] += 1
        else:
            monthly_direction[month_key]["unknown"] += 1
            direction_by_month_service[(month_key, service)]["unknown"] += 1

        if created:
            hourly_counter[created.strftime("%H:00")] += 1
            daily_counter[created.strftime("%Y-%m-%d")] += 1

    peak_hour_label, peak_hour_calls = ("--", 0)
    if hourly_counter:
        peak_hour_label, peak_hour_calls = max(hourly_counter.items(), key=lambda item: item[1])

    busiest_agent = next(iter(sorted(agent_totals.items(), key=lambda item: item[1]["calls"], reverse=True)), None)
    top_channel = trend_channel_counter.most_common(1)[0] if trend_channel_counter else ("Unknown Channel", 0)
    top_service_overall = service_counter.most_common(1)[0] if service_counter else ("Unspecified", 0)
    inbound_calls = call_type_counter["inbound"]
    outbound_calls = call_type_counter["outbound"]
    unknown_type_calls = call_type_counter["unknown"]

    window_label = "Sample snapshot"
    if created_values:
        minimum = min(created_values)
        maximum = max(created_values)
        if minimum.date() == maximum.date():
            window_label = f"{minimum:%d-%m-%Y} {minimum:%H:%M} to {maximum:%H:%M}"
        else:
            window_label = f"{minimum:%d-%m-%Y %H:%M} to {maximum:%d-%m-%Y %H:%M}"

    avg_handling = round(sum(handling_minutes) / len(handling_minutes), 1) if handling_minutes else 0.0
    max_handling = round(max(handling_minutes), 1) if handling_minutes else 0.0
    answer_rate = answered_calls / total_calls if total_calls else 0.0
    known_month_count = len([key for key in monthly_counter if key != UNKNOWN_PERIOD_KEY])
    single_month_dataset = known_month_count <= 1
    avg_monthly_volume = round(total_calls / known_month_count) if known_month_count else total_calls
    sorted_monthly_items = sorted(monthly_counter.items(), key=lambda item: (item[0] == UNKNOWN_PERIOD_KEY, item[0]))

    monthly_volume_series = [
        {
            "label": period_meta_from_key(key)["label"],
            "shortLabel": period_meta_from_key(key)["shortLabel"],
            "value": value,
            "isUnknown": period_meta_from_key(key)["isUnknown"],
        }
        for key, value in sorted_monthly_items
    ]
    channel_series = [{"label": label, "value": value} for label, value in channel_counter.most_common(8)]
    trend_channel_mix = build_display_channel_mix(trend_channel_counter)
    trend_channel_total = sum(item["value"] for item in trend_channel_mix)
    trend_channel_mix_valid = bool(receiving_channel_header) and trend_channel_total == total_calls
    trend_channel_warning_message = ""
    if not receiving_channel_header:
        trend_channel_warning_message = (
            "The uploaded dataset is missing the Receiving Channel column required for the channel breakdown section."
        )
    elif trend_channel_total != total_calls:
        trend_channel_warning_message = (
            "Channel totals do not match total inquiries in the uploaded dataset. Please review the Receiving Channel values."
        )
    service_series = [{"label": label, "value": value} for label, value in service_counter.most_common(8)]
    direction_series = [
        {
            "label": period_meta_from_key(key)["label"],
            "shortLabel": period_meta_from_key(key)["shortLabel"],
            "inbound": values["inbound"],
            "outbound": values["outbound"],
            "unknown": values["unknown"],
            "isUnknown": period_meta_from_key(key)["isUnknown"],
        }
        for key, values in sorted(monthly_direction.items(), key=lambda item: (item[0] == UNKNOWN_PERIOD_KEY, item[0]))
    ]
    direction_records = []
    for (month_key, service_name), values in sorted(
        direction_by_month_service.items(),
        key=lambda item: (item[0][0] == UNKNOWN_PERIOD_KEY, item[0][0], item[0][1]),
    ):
        month_meta = period_meta_from_key(month_key)
        month_dt = month_meta.get("datetime")
        direction_records.append(
            {
                "monthKey": month_key,
                "monthLabel": month_meta["label"],
                "monthShortLabel": month_meta["shortLabel"],
                "quarterKey": f"{month_dt.year}-Q{((month_dt.month - 1) // 3) + 1}" if month_dt else "",
                "quarterLabel": f"Q{((month_dt.month - 1) // 3) + 1} {month_dt.year}" if month_dt else "Unknown",
                "quarterShortLabel": f"Q{((month_dt.month - 1) // 3) + 1}" if month_dt else "Unknown",
                "yearKey": str(month_dt.year) if month_dt else "",
                "yearLabel": str(month_dt.year) if month_dt else "Unknown",
                "service": service_name,
                "inbound": values["inbound"],
                "outbound": values["outbound"],
                "unknown": values["unknown"],
                "isUnknownPeriod": month_meta["isUnknown"],
            }
        )
    direction_dataset_totals = {
        "all": {
            "inbound": inbound_calls,
            "outbound": outbound_calls,
            "unknown": unknown_type_calls,
            "total": total_calls,
        },
        "byService": {
            service_name: {
                "inbound": values["inbound"],
                "outbound": values["outbound"],
                "unknown": values["unknown"],
                "total": values["total"],
            }
            for service_name, values in sorted(direction_totals_by_service.items(), key=lambda item: item[0])
        },
    }
    top_agent_series = [
        {"label": agent, "value": metrics["calls"]}
        for agent, metrics in sorted(agent_totals.items(), key=lambda item: item[1]["calls"], reverse=True)[:10]
    ]
    busiest_period_source = [item for item in monthly_volume_series if not item.get("isUnknown")] or monthly_volume_series
    busiest_period = max(busiest_period_source, key=lambda item: item["value"]) if busiest_period_source else None
    unknown_period_total = next((item["value"] for item in monthly_volume_series if item.get("isUnknown")), 0)
    service_options = [label for label, _ in service_counter.most_common()]
    categorized_inquiries = sum(category_counter.values())
    categorization_rate = categorized_inquiries / total_calls if total_calls else 0.0
    category_summary_rows = [
        {
            "category": label,
            "count": value,
            "shareOfTotal": round(value / total_calls, 6) if total_calls else 0.0,
            "isUncategorized": False,
        }
        for label, value in category_counter.most_common()
    ]
    if uncategorized_inquiries:
        category_summary_rows.append(
            {
                "category": "Uncategorized",
                "count": uncategorized_inquiries,
                "shareOfTotal": round(uncategorized_inquiries / total_calls, 6) if total_calls else 0.0,
                "isUncategorized": True,
            }
        )
    category_options = [
        {
            "value": label,
            "label": label,
            "count": value,
        }
        for label, value in category_counter.most_common()
    ]

    analytics_variants = [
        {
            "key": "channel",
            "label": "Customer by channel",
            "chartType": "bar-list",
            "chartTitle": "Customer by channel",
            "chartNote": "Top receiving channels ranked by total customer volume.",
            "unitLabel": "customers",
            "series": channel_series,
        },
        {
            "key": "service",
            "label": "Customer by service",
            "chartType": "bar-list",
            "chartTitle": "Customer by service",
            "chartNote": "Most requested services in the current dataset by customer volume.",
            "unitLabel": "customers",
            "series": service_series,
        },
        {
            "key": "direction-month",
            "label": "Inbound vs outbound by month",
            "chartType": "direction-bi",
            "chartTitle": "Inbound vs outbound by month",
            "chartNote": "Month-over-month direction split from the uploaded dataset.",
            "series": direction_series,
            "records": direction_records,
            "datasetTotals": direction_dataset_totals,
            "serviceOptions": service_options,
            "yAxisLabel": "Number of Customers",
            "annotationThreshold": 0.2,
        },
        {
            "key": "agents",
            "label": "Top 10 agents by volume",
            "chartType": "bar-list",
            "chartTitle": "Top 10 agents by volume",
            "chartNote": "Customer-volume ranking for the busiest agents in the dataset.",
            "unitLabel": "customers",
            "series": top_agent_series,
        },
    ]

    table_rows = []
    for agent, metrics in sorted(agent_totals.items(), key=lambda item: item[1]["calls"], reverse=True):
        top_service = agent_service_counter[agent].most_common(1)[0][0] if agent_service_counter[agent] else "Unspecified"
        table_rows.append(
            {
                "agent": agent,
                "calls": metrics["calls"],
                "answerRate": round(metrics["answered"] / metrics["calls"], 4) if metrics["calls"] else 0,
                "inbound": metrics["inbound"],
                "outbound": metrics["outbound"],
                "avgHandlingMinutes": round(sum(metrics["handling"]) / len(metrics["handling"]), 1)
                if metrics["handling"]
                else 0.0,
                "topService": top_service,
            }
        )

    assumptions = [
        "Source metrics are seeded from the existing workbook snapshot rather than a live call-center API.",
    ]
    if len(unique_dates) <= 1:
        assumptions.append(
            "The workbook currently covers a single day, so the trend section is rendered as hourly volume instead of a multi-day trend."
        )
    if missed_calls == 0:
        assumptions.append(
            "Missed and abandoned customer KPIs look under-represented in this export, so answer-rate cards should be treated as sample values."
        )

    return {
        "modeKey": "general",
        "available": True,
        "meta": {
            "eyebrow": "Customer service center dashboard",
            "title": "Canadia Bank Customer service Center Dashboard",
            "subtitle": "Executive Summary of Service Center Operations.",
            "generatedAtIso": datetime.now().isoformat(timespec="seconds"),
            "generatedAtLabel": datetime.now().strftime("%d-%m-%Y %H:%M"),
            "windowLabel": window_label,
            "sourceWorkbook": source_label,
            "assumptions": assumptions,
            "todos": [
                "Swap ./data/call-center-dashboard.json for a live API or scheduled export when production metrics are ready.",
                "Add SLA, abandonment, and queue-level measures once those fields are present in the source workbook or API.",
            ],
        },
        "kpis": [
            {
                "id": "total-calls",
                "label": "Total Customers",
                "value": total_calls,
                "valueLabel": f"{total_calls:,}",
                "contextLabel": f"Snapshot window: {window_label}",
                "tone": "neutral",
            },
            {
                "id": "answer-rate",
                "label": "Answer Rate",
                "value": answer_rate,
                "valueLabel": f"{round(answer_rate * 100)}%",
                "contextLabel": f"{answered_calls} answered / {missed_calls} missed",
                "tone": "good",
            },
            {
                "id": "avg-handling",
                "label": "Avg Handling Time",
                "value": avg_handling,
                "valueLabel": f"{avg_handling:.1f} min",
                "contextLabel": f"Max observed: {max_handling:.1f} min",
                "tone": "neutral",
            },
            {
                "id": "peak-hour",
                "label": "Peak Hour",
                "value": peak_hour_calls,
                "valueLabel": peak_hour_label,
                "contextLabel": f"{peak_hour_calls} customers in busiest hour",
                "tone": "accent",
            },
        ],
        "sectionHeadings": {
            "kpis": {
                "eyebrow": "Headline KPIs",
                "title": "Operational pulse",
                "copy": "Top-line customer service center health for the current workbook snapshot.",
            },
            "trends": {
                "eyebrow": "Trends",
                "title": "Volume and mix",
                "copy": "Trend cards are intentionally thin so they can be swapped to live API series later.",
            },
            "category": {
                "eyebrow": "Category Summary",
                "title": "Inquiry categorization",
                "copy": "Review category coverage, isolate uncategorized inquiries, and keep the category summary management-friendly.",
            },
        },
        "analytics": {
            "eyebrow": "Operational analytics",
            "title": "Switch between key workload views",
            "description": "Use the chart buttons to compare channel mix, service demand, monthly direction split, and the busiest agents from the current dataset.",
            "defaultKey": "direction-month",
            "variants": analytics_variants,
            "highlights": [
                {
                    "label": "Busiest period",
                    "value": busiest_period["label"] if busiest_period else "--",
                    "context": (
                        f"{busiest_period['value']:,} customers (+{unknown_period_total:,} with unknown period)"
                        if busiest_period and unknown_period_total and not busiest_period.get("isUnknown")
                        else f"{busiest_period['value']:,} customers"
                    )
                    if busiest_period
                    else "No data available",
                },
                {
                    "label": "Average monthly volume",
                    "value": f"{avg_monthly_volume if not single_month_dataset else total_calls:,}",
                    "context": "Average = total customers (1 month only)"
                    if single_month_dataset
                    else "Average customers per month in the current dataset",
                },
                {
                    "label": "Top channel",
                    "value": top_channel[0],
                    "context": f"{top_channel[1]:,} customers",
                },
                {
                    "label": "Inbound share",
                    "value": f"{round((inbound_calls / total_calls) * 100) if total_calls else 0}%",
                    "context": "Share of inbound customers in the current dataset",
                },
                {
                    "label": "Top service",
                    "value": top_service_overall[0],
                    "context": f"{top_service_overall[1]:,} customers",
                },
                {
                    "label": "Top agent",
                    "value": busiest_agent[0] if busiest_agent else "--",
                    "context": f"{busiest_agent[1]['calls']:,} handled customers" if busiest_agent else "No data available",
                },
            ],
        },
        "trends": {
            "hourlyVolume": [
                {"label": label, "value": value} for label, value in sorted(hourly_counter.items(), key=lambda item: item[0])
            ],
            "serviceMix": [{"label": label, "value": value} for label, value in service_counter.most_common(5)],
            "channelMix": trend_channel_mix,
            "channelMixMeta": {
                "fieldLabel": receiving_channel_header or "Receiving Channel",
                "displayedChannelCount": len(trend_channel_mix),
                "displayedTotal": trend_channel_total,
                "totalInquiries": total_calls,
                "isValid": trend_channel_mix_valid,
                "warningMessage": trend_channel_warning_message,
            },
        },
        "categoryAnalysis": {
            "eyebrow": "Category summary",
            "title": "Inquiry categorization",
            "description": "Group inquiry records by category, keep uncategorized inquiries visible, and isolate missing category values when cleanup is needed.",
            "available": bool(category_header),
            "message": ""
            if category_header
            else "Inquiry category data is not available in the current dataset.",
            "categoryFieldLabel": category_header or "INQUIRY CATEGORY",
            "defaultStatus": "all",
            "defaultCategory": "all",
            "statusOptions": [
                {"value": "all", "label": "All"},
                {"value": "categorized-only", "label": "Categorized Only"},
                {"value": "uncategorized-only", "label": "Uncategorized Only"},
            ],
            "categoryOptions": category_options,
            "rows": category_summary_rows,
            "totals": {
                "totalInquiries": total_calls,
                "categorizedInquiries": categorized_inquiries,
                "uncategorizedInquiries": uncategorized_inquiries,
                "categorizationRate": round(categorization_rate, 6),
            },
        },
        "table": {
            "type": "agent-breakdown",
            "title": "Agent Breakdown",
            "description": "Sorted by handled customer volume so supervisors can expand this into SLA or coaching views later.",
            "rows": table_rows,
        },
    }


def build_window_label(created_values: list[datetime], fallback_label: str) -> str:
    if not created_values:
        return fallback_label
    minimum = min(created_values)
    maximum = max(created_values)
    if minimum.date() == maximum.date():
        return f"{minimum:%d-%m-%Y} {minimum:%H:%M} to {maximum:%H:%M}"
    return f"{minimum:%d-%m-%Y %H:%M} to {maximum:%d-%m-%Y %H:%M}"


def build_counter_series(counter: Counter[str], *, include_unknown: bool = True) -> list[dict[str, Any]]:
    items = counter.most_common()
    if not include_unknown:
        items = [(label, value) for label, value in items if label != "Unknown"]
    return [{"label": label, "value": value} for label, value in items]


def extract_complaint_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    complaint_rows = []
    for row in rows:
        if any(has_dimension_value(row.get(candidate)) for candidate in COMPLAINT_SIGNAL_CANDIDATES):
            complaint_rows.append(row)
    return complaint_rows


def build_complaint_payload(rows: list[dict[str, Any]], source_label: str) -> dict[str, Any]:
    headers = list(rows[0].keys()) if rows else []
    created_header = matching_header(headers, COMPLAINT_CREATED_CANDIDATES)
    category_header = matching_header(headers, COMPLAINT_CATEGORY_CANDIDATES)
    channel_header = matching_header(headers, COMPLAINT_CHANNEL_CANDIDATES)
    status_header = matching_header(headers, COMPLAINT_STATUS_CANDIDATES)
    detail_header = matching_header(headers, COMPLAINT_DETAIL_CANDIDATES)
    side_issue_header = matching_header(headers, COMPLAINT_SIDE_ISSUE_CANDIDATES)
    level_header = matching_header(headers, COMPLAINT_LEVEL_CANDIDATES)
    owner_header = matching_header(headers, COMPLAINT_OWNER_CANDIDATES)
    issue_header = matching_header(headers, ("Issue", "Complaint Issue"))

    complaint_rows = extract_complaint_rows(rows)

    created_values = [
        as_datetime(row.get(created_header or ""))
        for row in complaint_rows
        if created_header and as_datetime(row.get(created_header or "")) is not None
    ]

    total_complaints = len(complaint_rows)
    category_counter: Counter[str] = Counter()
    channel_counter: Counter[str] = Counter()
    status_counter: Counter[str] = Counter()
    detail_counter: Counter[str] = Counter()
    side_issue_counter: Counter[str] = Counter()
    level_counter: Counter[str] = Counter()
    owner_counter: Counter[str] = Counter()
    issue_counter: Counter[str] = Counter()
    monthly_counter: Counter[str] = Counter()
    daily_counter: Counter[str] = Counter()

    for row in complaint_rows:
        category_counter[normalize_dimension_label(row.get(category_header), fallback="Unknown")] += 1
        channel_counter[normalize_dimension_label(row.get(channel_header), fallback="Unknown")] += 1
        if status_header:
            status_counter[normalize_dimension_label(row.get(status_header), fallback="Unknown")] += 1
        if detail_header:
            detail_counter[normalize_dimension_label(row.get(detail_header), fallback="Unknown")] += 1
        if side_issue_header:
            side_issue_counter[normalize_side_issue_label(row.get(side_issue_header))] += 1
        if level_header:
            level_counter[normalize_dimension_label(row.get(level_header), fallback="Unknown")] += 1
        if owner_header:
            owner_counter[normalize_dimension_label(row.get(owner_header), fallback="Unknown")] += 1
        if issue_header:
            issue_counter[normalize_dimension_label(row.get(issue_header), fallback="Unknown")] += 1
        created = as_datetime(row.get(created_header or "")) if created_header else None
        if created:
            monthly_counter[created.strftime("%Y-%m")] += 1
            daily_counter[created.strftime("%Y-%m-%d")] += 1

    top_category = category_counter.most_common(1)[0] if category_counter else ("Unknown", 0)
    top_channel = channel_counter.most_common(1)[0] if channel_counter else ("Unknown", 0)
    top_status = status_counter.most_common(1)[0] if status_counter else None
    top_detail = detail_counter.most_common(1)[0] if detail_counter else None
    top_side_issue = side_issue_counter.most_common(1)[0] if side_issue_counter else None
    top_owner = owner_counter.most_common(1)[0] if owner_counter else None
    busiest_month = max(monthly_counter.items(), key=lambda item: item[1]) if monthly_counter else None
    category_total = sum(category_counter.values())
    channel_total = sum(channel_counter.values())
    status_total = sum(status_counter.values()) if status_header else 0
    detail_total = sum(detail_counter.values()) if detail_header else 0
    side_issue_total = sum(side_issue_counter.values()) if side_issue_header else 0
    level_total = sum(level_counter.values()) if level_header else 0
    owner_total = sum(owner_counter.values()) if owner_header else 0
    known_category_total = sum(value for label, value in category_counter.items() if label != "Unknown")
    known_channel_total = sum(value for label, value in channel_counter.items() if label != "Unknown")
    known_status_total = sum(value for label, value in status_counter.items() if label != "Unknown") if status_header else 0
    known_detail_total = sum(value for label, value in detail_counter.items() if label != "Unknown") if detail_header else 0
    known_side_issue_total = sum(value for label, value in side_issue_counter.items() if label != "Unknown") if side_issue_header else 0
    known_level_total = sum(value for label, value in level_counter.items() if label != "Unknown") if level_header else 0
    known_owner_total = sum(value for label, value in owner_counter.items() if label != "Unknown") if owner_header else 0

    category_series = build_counter_series(category_counter)
    channel_series = build_counter_series(channel_counter)
    status_series = build_counter_series(status_counter)
    detail_series = build_counter_series(detail_counter)
    side_issue_series = build_counter_series(side_issue_counter)
    level_series = build_counter_series(level_counter)
    owner_series = build_counter_series(owner_counter)
    issue_series = build_counter_series(issue_counter)
    monthly_series = [
        {
            "label": datetime.strptime(key, "%Y-%m").strftime("%b %Y"),
            "shortLabel": datetime.strptime(key, "%Y-%m").strftime("%b"),
            "value": value,
        }
        for key, value in sorted(monthly_counter.items(), key=lambda item: item[0])
    ]
    table_rows = [
        {
            "category": label,
            "count": value,
            "shareOfTotal": round(value / total_complaints, 6) if total_complaints else 0.0,
        }
        for label, value in category_counter.most_common()
    ]

    analytics_variants = [
        {
            "key": "complaint-category",
            "label": "Complaint by category",
            "chartType": "bar-list",
            "chartTitle": "Complaint count by category",
            "chartNote": f"Detected from {category_header or 'the complaint category field'} in the uploaded complaint dataset.",
            "unitLabel": "complaints",
            "series": category_series,
        },
        {
            "key": "complaint-channel",
            "label": "Complaint by channel",
            "chartType": "bar-list",
            "chartTitle": "Complaint count by channel",
            "chartNote": f"Detected from {channel_header or 'the complaint channel field'} in the uploaded complaint dataset.",
            "unitLabel": "complaints",
            "series": channel_series,
        },
    ]
    if status_header:
        analytics_variants.append(
            {
                "key": "complaint-status",
            "label": "Complaint by status",
            "chartType": "bar-list",
            "chartTitle": "Complaint count by status",
            "chartNote": f"Detected from {status_header} in the uploaded complaint dataset.",
            "unitLabel": "complaints",
            "series": status_series,
            }
        )
    if monthly_series:
        analytics_variants.append(
            {
                "key": "complaint-month",
                "label": "Complaint by month",
                "chartType": "column",
                "chartTitle": "Complaint count by month",
                "chartNote": f"Detected from {created_header} in the uploaded complaint dataset.",
                "unitLabel": "complaints",
                "series": monthly_series,
            }
        )
    if detail_header:
        analytics_variants.append(
            {
                "key": "complaint-detail",
                "label": "Complaint by detail",
                "chartType": "bar-list",
                "chartTitle": "Complaint count by detail",
                "chartNote": f"Detected from {detail_header} in the uploaded complaint dataset.",
                "unitLabel": "complaints",
                "series": detail_series,
            }
        )
    if side_issue_header:
        analytics_variants.append(
            {
                "key": "complaint-side-issue",
                "label": "Complaint by side issue",
                "chartType": "bar-list",
                "chartTitle": "Complaint count by side issue",
                "chartNote": f"Detected from {side_issue_header} in the uploaded complaint dataset.",
                "unitLabel": "complaints",
                "series": side_issue_series,
            }
        )
    if level_header:
        analytics_variants.append(
            {
                "key": "complaint-level",
                "label": "Complaint by level",
                "chartType": "bar-list",
                "chartTitle": "Complaint count by level",
                "chartNote": f"Detected from {level_header} in the uploaded complaint dataset.",
                "unitLabel": "complaints",
                "series": level_series,
            }
        )
    if owner_header:
        analytics_variants.append(
            {
                "key": "complaint-owner",
                "label": "Complaint by owner",
                "chartType": "bar-list",
                "chartTitle": "Complaint count by business owner",
                "chartNote": f"Detected from {owner_header} in the uploaded complaint dataset.",
                "unitLabel": "complaints",
                "series": owner_series,
            }
        )

    return {
        "modeKey": "complaint",
        "available": True,
        "meta": {
            "eyebrow": "Complaint dashboard",
            "title": "Complaint Dashboard",
            "subtitle": "Management view of complaint intake, channels, and resolution status.",
            "generatedAtIso": datetime.now().isoformat(timespec="seconds"),
            "generatedAtLabel": datetime.now().strftime("%d-%m-%Y %H:%M"),
            "windowLabel": build_window_label(created_values, "Complaint dataset snapshot"),
            "sourceWorkbook": source_label,
            "assumptions": [
                "Complaint mode is built from the Complaints sheet only and keeps General mode separate.",
                "Summary sheet is treated as a reference for which metrics matter, but all complaint calculations come from the raw Complaints sheet.",
                "Blank, null, and whitespace-only complaint fields are grouped under Unknown so totals always reconcile.",
            ],
            "todos": [],
        },
        "kpis": [
            {
                "id": "total-complaints",
                "label": "Total Complaints",
                "value": total_complaints,
                "valueLabel": f"{total_complaints:,}",
                "contextLabel": "Complaint-bearing rows from the raw Complaints sheet",
                "tone": "neutral",
            },
            {
                "id": "complaint-categories",
                "label": "Complaint Categories",
                "value": len(category_counter),
                "valueLabel": f"{len(category_counter):,}",
                "contextLabel": f"Detected from {category_header or 'Unknown'} in Complaints",
                "tone": "neutral",
            },
            {
                "id": "complaint-channels",
                "label": "Complaint Channels",
                "value": len(channel_counter),
                "valueLabel": f"{len(channel_counter):,}",
                "contextLabel": f"Detected from {channel_header or 'Unknown'} in Complaints",
                "tone": "neutral",
            },
            {
                "id": "complaint-statuses",
                "label": "Complaint Statuses",
                "value": len(status_counter) if status_header else 0,
                "valueLabel": f"{len(status_counter):,}" if status_header else "--",
                "contextLabel": f"Detected from {status_header}" if status_header else "Status column not detected",
                "tone": "neutral",
            },
            {
                "id": "top-complaint-category",
                "label": "Top Complaint Category",
                "value": top_category[1],
                "valueLabel": top_category[0],
                "contextLabel": f"{top_category[1]:,} complaints",
                "tone": "accent",
            },
            {
                "id": "top-complaint-channel",
                "label": "Top Complaint Channel",
                "value": top_channel[1],
                "valueLabel": top_channel[0],
                "contextLabel": f"{top_channel[1]:,} complaints",
                "tone": "accent",
            },
        ],
        "sectionHeadings": {
            "kpis": {
                "eyebrow": "Headline KPIs",
                "title": "Complaint overview",
                "copy": "Top-line complaint intake rebuilt from the raw Complaints sheet.",
            },
            "trends": {
                "eyebrow": "Complaint views",
                "title": "Category and channel distribution",
                "copy": "Compare complaint mix, monthly trend, and Summary-style support metrics using the same dashboard style as General mode.",
            },
            "category": {
                "eyebrow": "Status Summary",
                "title": "Complaint status overview",
                "copy": "Review complaint status, severity, and ownership metrics while keeping all totals reconciled to the raw Complaints sheet.",
            },
        },
        "analytics": {
            "eyebrow": "Complaint analytics",
            "title": "Switch between complaint views",
            "description": "Use the chart buttons to compare complaint category, channel, status, detail, side issue, severity, ownership, and monthly volume from the raw Complaints sheet.",
            "defaultKey": "complaint-category",
            "variants": analytics_variants,
            "highlights": [
                {
                    "label": "Top complaint category",
                    "value": top_category[0],
                    "context": f"{top_category[1]:,} complaints",
                },
                {
                    "label": "Top complaint channel",
                    "value": top_channel[0],
                    "context": f"{top_channel[1]:,} complaints",
                },
                {
                    "label": "Top complaint status",
                    "value": top_status[0] if top_status else "--",
                    "context": f"{top_status[1]:,} complaints" if top_status else "Status column not detected",
                },
                {
                    "label": "Top side issue",
                    "value": top_side_issue[0] if top_side_issue else "--",
                    "context": f"{top_side_issue[1]:,} complaints" if top_side_issue else "Side issue column not detected",
                },
                {
                    "label": "Top business owner",
                    "value": top_owner[0] if top_owner else "--",
                    "context": f"{top_owner[1]:,} complaints" if top_owner else "Business owner column not detected",
                },
                {
                    "label": "Busiest month",
                    "value": datetime.strptime(busiest_month[0], "%Y-%m").strftime("%b %Y") if busiest_month else "--",
                    "context": f"{busiest_month[1]:,} complaints" if busiest_month else "Created date column not detected",
                },
            ],
        },
        "trends": {
            "monthlyVolume": monthly_series,
            "categoryBreakdown": category_series,
            "channelBreakdown": channel_series,
            "statusBreakdown": status_series,
            "detailBreakdown": detail_series,
            "sideIssueBreakdown": side_issue_series,
            "levelBreakdown": level_series,
            "ownerBreakdown": owner_series,
            "issueBreakdown": issue_series,
            "totals": {
                "totalComplaints": total_complaints,
                "categoryTotal": category_total,
                "channelTotal": channel_total,
                "statusTotal": status_total,
                "detailTotal": detail_total,
                "sideIssueTotal": side_issue_total,
                "levelTotal": level_total,
                "ownerTotal": owner_total,
            },
        },
        "categoryAnalysis": {
            "available": bool(status_header or level_header or owner_header),
            "title": "Complaint status overview",
            "description": "Status, complaint level, and business owner counts are grouped directly from the raw Complaints sheet and include Unknown when the source field is blank.",
            "statusFieldLabel": status_header or "Status",
            "rows": [
                {
                    "label": label,
                    "count": value,
                    "shareOfTotal": round(value / total_complaints, 6) if total_complaints else 0.0,
                }
                for label, value in status_counter.most_common()
            ],
            "totals": {
                "totalComplaints": total_complaints,
                "statusGroups": len(status_counter),
                "knownStatusComplaints": known_status_total,
                "unknownStatusComplaints": total_complaints - known_status_total if status_header else total_complaints,
                "levelGroups": len(level_counter),
                "knownLevelComplaints": known_level_total,
                "unknownLevelComplaints": total_complaints - known_level_total if level_header else total_complaints,
                "ownerGroups": len(owner_counter),
                "knownOwnerComplaints": known_owner_total,
                "unknownOwnerComplaints": total_complaints - known_owner_total if owner_header else total_complaints,
            },
            "levelFieldLabel": level_header or "Complaint Level",
            "levelRows": [
                {
                    "label": label,
                    "count": value,
                    "shareOfTotal": round(value / total_complaints, 6) if total_complaints else 0.0,
                }
                for label, value in level_counter.most_common()
            ],
            "ownerFieldLabel": owner_header or "Business Owner",
            "ownerRows": [
                {
                    "label": label,
                    "count": value,
                    "shareOfTotal": round(value / total_complaints, 6) if total_complaints else 0.0,
                }
                for label, value in owner_counter.most_common()
            ],
            "message": "Status, complaint level, and business owner columns were not detected in the uploaded complaint dataset."
            if not (status_header or level_header or owner_header)
            else "",
        },
        "table": {
            "type": "complaint-breakdown",
            "title": "Complaint Breakdown",
            "description": "Complaint counts by category with percentage of total complaints.",
            "rows": table_rows,
        },
    }


def build_complaint_management_payload(rows: list[dict[str, Any]], source_label: str) -> dict[str, Any]:
    headers = list(rows[0].keys()) if rows else []
    created_header = matching_header(headers, COMPLAINT_CREATED_CANDIDATES)
    month_header = matching_header(headers, COMPLAINT_MONTH_CANDIDATES)
    issue_header = matching_header(headers, ("Issue", "Complaint Issue"))
    detail_header = matching_header(headers, COMPLAINT_DETAIL_CANDIDATES)
    category_header = matching_header(headers, COMPLAINT_CATEGORY_SOURCE_CANDIDATES)
    subcategory_header = matching_header(headers, COMPLAINT_SUBCATEGORY_CANDIDATES)
    status_header = matching_header(headers, COMPLAINT_STATUS_CANDIDATES)
    level_header = matching_header(headers, COMPLAINT_LEVEL_CANDIDATES)
    channel_header = matching_header(headers, COMPLAINT_CHANNEL_CANDIDATES)
    owner_header = matching_header(headers, COMPLAINT_OWNER_CANDIDATES)
    gender_header = matching_header(headers, COMPLAINT_GENDER_CANDIDATES)
    side_issue_header = matching_header(headers, COMPLAINT_SIDE_ISSUE_CANDIDATES)
    feedback_header = matching_header(headers, COMPLAINT_FEEDBACK_CANDIDATES)
    resolution_header = matching_header(headers, COMPLAINT_RESOLUTION_CANDIDATES)
    service_header = matching_header(headers, COMPLAINT_SERVICE_CANDIDATES)

    complaint_rows = extract_complaint_rows(rows)
    time_window_values: list[datetime] = []
    normalized_records: list[dict[str, Any]] = []

    issue_counter: Counter[str] = Counter()
    detail_counter: Counter[str] = Counter()
    status_counter: Counter[str] = Counter()
    level_counter: Counter[str] = Counter()
    channel_counter: Counter[str] = Counter()
    owner_counter: Counter[str] = Counter()
    gender_counter: Counter[str] = Counter()
    side_issue_counter: Counter[str] = Counter()
    feedback_counter: Counter[str] = Counter()
    trend_counter: Counter[str] = Counter()
    taxonomy_fallback_count = 0
    feedback_data_detected = False

    for index, row in enumerate(complaint_rows, start=1):
        raw_issue = normalize_dimension_label(row.get(issue_header), fallback="Unknown") if issue_header else "Unknown"
        raw_detail = normalize_dimension_label(row.get(detail_header), fallback="Unknown") if detail_header else "Unknown"
        raw_category = normalize_dimension_label(row.get(category_header), fallback="Unknown") if category_header else "Unknown"
        raw_subcategory = (
            normalize_dimension_label(row.get(subcategory_header), fallback="Unknown") if subcategory_header else "Unknown"
        )
        raw_service = normalize_dimension_label(row.get(service_header), fallback="Unknown") if service_header else "Unknown"
        issue, detail, used_taxonomy_fallback = map_complaint_record_to_taxonomy(
            issue_value=row.get(issue_header) if issue_header else raw_issue,
            detail_value=row.get(detail_header) if detail_header else raw_detail,
            category_value=row.get(category_header) if category_header else raw_category,
            subcategory_value=row.get(subcategory_header) if subcategory_header else raw_subcategory,
            service_value=row.get(service_header) if service_header else raw_service,
        )
        status = normalize_status_label(row.get(status_header)) if status_header else "Unknown"
        level = normalize_level_label(row.get(level_header)) if level_header else "Unknown"
        channel = normalize_dimension_label(row.get(channel_header), fallback="Unknown") if channel_header else "Unknown"
        owner = normalize_dimension_label(row.get(owner_header), fallback="Unknown") if owner_header else "Unknown"
        gender = normalize_gender_label(row.get(gender_header)) if gender_header else "Unknown"
        side_issue = normalize_side_issue_label(row.get(side_issue_header)) if side_issue_header else "Unknown"
        side_issue_group = normalize_matrix_side_issue_label(row.get(side_issue_header)) if side_issue_header else "Unknown"
        complaint_feedback = normalize_complaint_feedback_label(row.get(feedback_header)) if feedback_header else "Unknown"
        resolution = clean_text(row.get(resolution_header)) if resolution_header else ""
        if complaint_feedback != "Unknown":
            feedback_data_detected = True
        if used_taxonomy_fallback:
            taxonomy_fallback_count += 1

        created = as_datetime(row.get(created_header or "")) if created_header else None
        month_key = created.strftime("%Y-%m") if created else as_month_key(row.get(month_header or "")) if month_header else None
        period_key = month_key or UNKNOWN_PERIOD_KEY
        period_meta = period_meta_from_key(period_key)
        if created:
            time_window_values.append(created)
        elif month_key:
            time_window_values.append(datetime.strptime(month_key, "%Y-%m"))

        normalized_records.append(
            {
                "id": index,
                "issue": issue,
                "detail": detail,
                "rawIssue": raw_issue,
                "rawDetail": raw_detail,
                "rawCategory": raw_category,
                "rawSubCategory": raw_subcategory,
                "rawService": raw_service,
                "status": status,
                "level": level,
                "channel": channel,
                "owner": owner,
                "gender": gender,
                "sideIssue": side_issue,
                "sideIssueGroup": side_issue_group,
                "complaintFeedback": complaint_feedback,
                "periodKey": period_key,
                "periodLabel": period_meta["label"],
                "periodShortLabel": period_meta["shortLabel"],
                "isUnknownPeriod": period_meta["isUnknown"],
                "createdIso": created.isoformat() if created else "",
                "createdLabel": created.strftime("%d-%b-%Y %H:%M") if created else "",
                "resolution": resolution,
                "usedTaxonomyFallback": used_taxonomy_fallback,
            }
        )

        issue_counter[issue] += 1
        detail_counter[detail] += 1
        status_counter[status] += 1
        level_counter[level] += 1
        channel_counter[channel] += 1
        owner_counter[owner] += 1
        gender_counter[gender] += 1
        side_issue_counter[side_issue_group] += 1
        feedback_counter[complaint_feedback] += 1
        trend_counter[period_key] += 1

    total_complaints = len(normalized_records)
    feedback_kpi_counter = Counter(feedback_counter)
    if not feedback_header or not feedback_data_detected:
        feedback_kpi_counter = Counter({"Complaint": total_complaints, "Feedback": 0, "Unknown": 0})

    issue_series = [
        {
            "label": issue_name,
            "value": issue_counter.get(issue_name, 0),
        }
        for issue_name in COMPLAINT_ALL_ISSUES
    ]
    channel_series = build_counter_series(channel_counter)
    owner_series = build_counter_series(owner_counter)
    trend_series = [
        {
            "label": period_meta_from_key(key)["label"],
            "shortLabel": period_meta_from_key(key)["shortLabel"],
            "value": value,
            "isUnknown": key == UNKNOWN_PERIOD_KEY,
        }
        for key, value in sorted(trend_counter.items(), key=lambda item: (item[0] == UNKNOWN_PERIOD_KEY, item[0]))
    ]

    missing_metrics: list[str] = []
    if not issue_header:
        missing_metrics.append("Issue column was not detected. Main Issue totals are mapped from the available complaint fields.")
    if not detail_header:
        missing_metrics.append("Breakdown column was not detected. Detail totals are mapped from the available complaint fields.")
    if not status_header:
        missing_metrics.append("Status metrics are unavailable because the Status column was not detected.")
    if not level_header:
        missing_metrics.append("Complaint level metrics are unavailable because the Complaint Level column was not detected.")
    if not channel_header:
        missing_metrics.append("Channel metrics are unavailable because the Receiving Channels column was not detected.")
    if not owner_header:
        missing_metrics.append("Business owner metrics are unavailable because the Business Owner column was not detected.")
    if not gender_header:
        missing_metrics.append("Gender summary is unavailable because the Gender column was not detected.")
    if not side_issue_header:
        missing_metrics.append("Side-issue ownership columns are unavailable because the Side issue column was not detected.")
    if not feedback_header:
        missing_metrics.append(
            "Complaint/Feedback column was not detected. Complaint KPI treats all complaint rows as Complaint and Feedback as 0."
        )
    elif not feedback_data_detected:
        missing_metrics.append(
            "Complaint/Feedback is blank in the complaint rows. Complaint KPI treats all complaint rows as Complaint and Feedback as 0."
        )
    if taxonomy_fallback_count:
        missing_metrics.append(
            f"{taxonomy_fallback_count:,} complaint rows used fallback taxonomy mapping so every complaint stays inside the fixed management structure."
        )

    default_kpis = [
        ("Total Complaints", total_complaints, "Complaint-bearing rows from the raw Complaints sheet", "neutral"),
        ("Close", status_counter.get("Close", 0), "Count from Status" if status_header else "Status column not detected", "neutral"),
        ("Progress", status_counter.get("Progress", 0), "Count from Status" if status_header else "Status column not detected", "neutral"),
        (
            "Complaint",
            feedback_kpi_counter.get("Complaint", 0),
            "Derived from Complaint/Feedback" if feedback_header and feedback_data_detected else "All complaint rows treated as Complaint",
            "neutral",
        ),
        (
            "Feedback",
            feedback_kpi_counter.get("Feedback", 0),
            "Derived from Complaint/Feedback" if feedback_header and feedback_data_detected else "Data unavailable for this metric",
            "neutral",
        ),
        ("Critical", level_counter.get("Critical", 0), "Count from Complaint Level" if level_header else "Complaint Level column not detected", "accent"),
        ("High", level_counter.get("High", 0), "Count from Complaint Level" if level_header else "Complaint Level column not detected", "accent"),
        ("Medium", level_counter.get("Medium", 0), "Count from Complaint Level" if level_header else "Complaint Level column not detected", "accent"),
        ("Low", level_counter.get("Low", 0), "Count from Complaint Level" if level_header else "Complaint Level column not detected", "accent"),
    ]

    return {
        "modeKey": "complaint",
        "available": True,
        "meta": {
            "eyebrow": "Complaint dashboard",
            "title": "Complaint Dashboard",
            "subtitle": "Management view rebuilt dynamically from the raw Complaints sheet.",
            "generatedAtIso": datetime.now().isoformat(timespec="seconds"),
            "generatedAtLabel": datetime.now().strftime("%d-%m-%Y %H:%M"),
            "windowLabel": build_window_label(time_window_values, "Complaint dataset snapshot"),
            "sourceWorkbook": source_label,
            "assumptions": [
                "Complaint mode uses the raw Complaints sheet only; Summary is reference-only.",
                "Main Issue and Detail use a fixed management taxonomy while counts are recalculated dynamically from the uploaded complaint rows.",
                "Blank, null, and whitespace-only complaint dimensions are grouped under Unknown so totals always reconcile.",
            ],
            "todos": [],
        },
        "kpis": [
            {
                "id": f"complaint-kpi-{index}",
                "label": label,
                "value": value,
                "valueLabel": f"{value:,}",
                "contextLabel": context,
                "tone": tone,
            }
            for index, (label, value, context, tone) in enumerate(default_kpis, start=1)
        ],
        "sectionHeadings": {
            "kpis": {
                "eyebrow": "Headline KPIs",
                "title": "Complaint KPI snapshot",
                "copy": "Status, severity, and complaint-type cards are recalculated from the filtered raw Complaints rows.",
            },
            "trends": {
                "eyebrow": "Complaint charts",
                "title": "Volume and ownership views",
                "copy": "Visual summaries are rebuilt from the same filtered complaint scope so management can compare Main Issue, channel, owner, and trend quickly.",
            },
            "category": {
                "eyebrow": "Summary sections",
                "title": "Complaint summaries",
                "copy": "Compact summaries mirror the management themes from Summary while staying fully dynamic from raw Complaints rows.",
            },
        },
        "analytics": {
            "eyebrow": "Complaint filters",
            "title": "Refine complaint scope",
            "description": "Use the fixed Main Issue and Detail taxonomy to update KPI cards, the management matrix, summaries, charts, and detail records. All values come from the raw Complaints sheet.",
            "variants": [],
            "highlights": [],
        },
        "trends": {
            "issueBreakdown": issue_series,
            "channelBreakdown": channel_series,
            "ownerBreakdown": owner_series,
            "trendSeries": trend_series,
            "totals": {
                "totalComplaints": total_complaints,
                "issueTotal": sum(issue_counter.values()),
                "channelTotal": sum(channel_counter.values()),
                "ownerTotal": sum(owner_counter.values()),
                "trendTotal": sum(trend_counter.values()),
            },
        },
        "categoryAnalysis": {
            "available": True,
            "title": "Complaint summaries",
            "description": "Compact complaint summaries are grouped directly from the filtered Complaints rows.",
            "rows": [],
            "totals": {"totalComplaints": total_complaints},
            "message": "",
        },
        "table": {
            "type": "complaint-matrix",
            "title": "Complaint Matrix",
            "description": "Summary-style complaint matrix rebuilt dynamically from raw complaint rows.",
            "rows": [],
        },
        "complaintModule": {
            "fieldLabels": {
                "issue": issue_header or "Issue",
                "detail": detail_header or "Breakdown",
                "category": category_header or "Category",
                "subCategory": subcategory_header or "Sub Category",
                "status": status_header or "Status",
                "level": level_header or "Complaint Level",
                "channel": channel_header or "Receiving Channels",
                "owner": owner_header or "Business Owner",
                "gender": gender_header or "Gender",
                "sideIssue": side_issue_header or "Side issue",
                "complaintFeedback": feedback_header or "Complaint/Feedback",
                "service": service_header or "Service",
                "created": created_header or "Created",
                "month": month_header or "Month",
                "resolution": resolution_header or "Resolution",
            },
            "availability": {
                "issue": bool(issue_header or detail_header or category_header or subcategory_header or service_header),
                "detail": bool(detail_header or subcategory_header or category_header or issue_header or service_header),
                "status": bool(status_header),
                "level": bool(level_header),
                "channel": bool(channel_header),
                "owner": bool(owner_header),
                "gender": bool(gender_header),
                "sideIssue": bool(side_issue_header),
                "complaintFeedback": bool(feedback_header and feedback_data_detected),
                "time": bool(created_header or month_header),
                "resolution": bool(resolution_header),
            },
            "totals": {
                "totalComplaints": total_complaints,
                "issueTotal": sum(issue_counter.values()),
                "detailTotal": sum(detail_counter.values()),
                "statusTotal": sum(status_counter.values()),
                "levelTotal": sum(level_counter.values()),
                "channelTotal": sum(channel_counter.values()),
                "ownerTotal": sum(owner_counter.values()),
                "genderTotal": sum(gender_counter.values()),
                "sideIssueTotal": sum(side_issue_counter.values()),
                "complaintFeedbackTotal": sum(feedback_kpi_counter.values()),
            },
            "defaultFilters": {"issue": "all", "detail": "all"},
            "taxonomy": {
                "issues": [
                    {
                        "label": issue_name,
                        "details": details,
                    }
                    for issue_name, details in COMPLAINT_FIXED_TAXONOMY
                ]
            },
            "missingMetrics": missing_metrics,
            "records": normalized_records,
        },
    }


def build_unavailable_complaint_payload(source_label: str, error: DataValidationError | None = None) -> dict[str, Any]:
    message = error.public_message if error else "Complaint dataset is not available yet."
    details = error.details if error else ["Upload a valid complaint dataset to enable Complaint mode."]
    return {
        "modeKey": "complaint",
        "available": False,
        "meta": {
            "eyebrow": "Complaint dashboard",
            "title": "Complaint Dashboard",
            "subtitle": "Management view of complaint intake, channels, and resolution status.",
            "generatedAtIso": datetime.now().isoformat(timespec="seconds"),
            "generatedAtLabel": datetime.now().strftime("%d-%m-%Y %H:%M"),
            "windowLabel": "Complaint dataset unavailable",
            "sourceWorkbook": source_label,
            "assumptions": details,
            "todos": [],
        },
        "kpis": [],
        "sectionHeadings": {
            "kpis": {
                "eyebrow": "Headline KPIs",
                "title": "Complaint overview",
                "copy": "Upload a complaint dataset to enable complaint KPIs.",
            },
            "trends": {
                "eyebrow": "Complaint views",
                "title": "Category and channel distribution",
                "copy": "Complaint visuals appear once a valid complaint dataset is available.",
            },
            "category": {
                "eyebrow": "Status Summary",
                "title": "Complaint status overview",
                "copy": "Complaint status content appears once a valid complaint dataset is available.",
            },
        },
        "analytics": {
            "eyebrow": "Complaint analytics",
            "title": "Switch between complaint views",
            "description": message,
            "defaultKey": "complaint-category",
            "variants": [],
            "highlights": [],
        },
        "trends": {
            "categoryBreakdown": [],
            "channelBreakdown": [],
            "totals": {"totalComplaints": 0, "categoryTotal": 0, "channelTotal": 0, "statusTotal": 0},
        },
        "categoryAnalysis": {
            "available": False,
            "title": "Complaint status overview",
            "description": message,
            "rows": [],
            "totals": {"totalComplaints": 0, "statusGroups": 0, "knownStatusComplaints": 0, "unknownStatusComplaints": 0},
            "message": message,
        },
        "table": {
            "type": "complaint-breakdown",
            "title": "Complaint Breakdown",
            "description": message,
            "rows": [],
        },
    }


def write_payload(payload: dict[str, Any], output_path: Path = OUTPUT_PATH) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_suffix(".tmp")
    temp_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    temp_path.replace(output_path)


def build_dashboard_json(
    dataset_path: Path,
    *,
    output_path: Path = OUTPUT_PATH,
    source_label: str | None = None,
    write_json_file: bool = True,
) -> dict[str, Any]:
    rows = load_and_validate_rows(dataset_path, mode="general")
    payload = build_payload(rows, source_label or dataset_path.name)
    if write_json_file:
        write_payload(payload, output_path)
    return payload


def build_complaint_dashboard_json(
    dataset_path: Path,
    *,
    output_path: Path = OUTPUT_PATH,
    source_label: str | None = None,
    write_json_file: bool = True,
) -> dict[str, Any]:
    rows = load_and_validate_rows(dataset_path, mode="complaint")
    payload = build_complaint_management_payload(rows, source_label or dataset_path.name)
    if write_json_file:
        write_payload(payload, output_path)
    return payload


def get_active_dataset_state_path(mode: str) -> Path:
    return ACTIVE_DATASET_STATE_PATH if mode == "general" else ACTIVE_COMPLAINT_DATASET_STATE_PATH


def save_active_dataset(stored_path: Path, original_name: str, *, mode: str = "general") -> None:
    state_path = get_active_dataset_state_path(mode)
    state_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.write_text(
        json.dumps(
            {
                "storedPath": str(stored_path),
                "originalName": original_name,
                "updatedAt": datetime.now().isoformat(timespec="seconds"),
                "mode": mode,
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def load_active_dataset_state(mode: str = "general") -> dict[str, Any] | None:
    state_path = get_active_dataset_state_path(mode)
    if not state_path.exists():
        return None
    try:
        return json.loads(state_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None


def resolve_active_dataset(mode: str = "general") -> DatasetSource:
    state = load_active_dataset_state(mode)
    if state:
        stored_path = Path(str(state.get("storedPath", "")))
        original_name = clean_text(state.get("originalName")) or stored_path.name
        if stored_path.exists():
            return DatasetSource(path=stored_path, label=original_name)
    default_path = DEFAULT_DATASET_PATH if mode == "general" else DEFAULT_COMPLAINT_DATASET_PATH
    return DatasetSource(path=default_path, label=default_path.name)


def build_unavailable_general_payload(source_label: str, error: Exception | None = None) -> dict[str, Any]:
    details = error.details if isinstance(error, DataValidationError) else []
    reason = (
        error.public_message
        if isinstance(error, DataValidationError)
        else "Upload a General dataset to load this dashboard mode."
    )
    generated_at = datetime.now()
    follow_up = [reason, *details] if details else [reason]

    return {
        "modeKey": "general",
        "available": False,
        "meta": {
            "eyebrow": "Customer service center dashboard",
            "title": "Call Center Dashboard",
            "subtitle": "Upload a General dataset to load this dashboard mode.",
            "generatedAtIso": generated_at.isoformat(timespec="seconds"),
            "generatedAtLabel": generated_at.strftime("%d-%m-%Y %H:%M"),
            "windowLabel": "No dataset uploaded",
            "sourceWorkbook": source_label,
            "assumptions": [],
            "todos": follow_up,
        },
        "kpis": [],
        "sectionHeadings": {
            "kpis": {
                "eyebrow": "Headline KPIs",
                "title": "Operational pulse",
                "copy": "Upload a General dataset to populate KPI cards.",
            },
            "trends": {
                "eyebrow": "Trends",
                "title": "Volume and mix",
                "copy": "Upload a General dataset to populate charts and trend summaries.",
            },
            "category": {
                "eyebrow": "Category Summary",
                "title": "Inquiry categorization",
                "copy": "Upload a General dataset to populate inquiry category analytics.",
            },
        },
        "analytics": {
            "eyebrow": "Operational analytics",
            "title": "Upload a General dataset",
            "description": reason,
            "defaultKey": None,
            "variants": [],
        },
        "trends": {
            "cards": [],
            "serviceMix": [],
            "channelMix": [],
            "channelMixMeta": {
                "fieldLabel": "Receiving Channel",
                "displayedChannelCount": 0,
                "isValid": True,
                "warningMessage": "",
            },
        },
        "categoryAnalysis": {
            "available": False,
            "message": "Upload a General dataset to populate the inquiry categorization section.",
        },
        "table": {
            "available": False,
            "message": reason,
            "rows": [],
            "columns": [],
        },
    }


def build_dashboard_bundle(*, output_path: Path = OUTPUT_PATH, write_json_file: bool = True) -> dict[str, Any]:
    general_source = resolve_active_dataset("general")
    complaint_source = resolve_active_dataset("complaint")
    try:
        general_payload = build_dashboard_json(
            general_source.path,
            source_label=general_source.label,
            write_json_file=False,
        )
    except DataValidationError as error:
        general_payload = build_unavailable_general_payload(general_source.label, error)
    except FileNotFoundError:
        general_payload = build_unavailable_general_payload(general_source.label)
    try:
        complaint_payload = build_complaint_dashboard_json(
            complaint_source.path,
            source_label=complaint_source.label,
            write_json_file=False,
        )
    except DataValidationError as error:
        complaint_payload = build_unavailable_complaint_payload(complaint_source.label, error)
    except FileNotFoundError:
        complaint_payload = build_unavailable_complaint_payload(complaint_source.label)

    payload = {
        "defaultMode": "general",
        "modeOptions": [
            {"key": "general", "label": "General", "available": True},
            {"key": "complaint", "label": "Complaint", "available": complaint_payload.get("available", False)},
        ],
        "modes": {
            "general": general_payload,
            "complaint": complaint_payload,
        },
    }
    if write_json_file:
        write_payload(payload, output_path)
    return payload


def activate_uploaded_dataset(temp_path: Path, original_name: str, *, mode: str = "general") -> dict[str, Any]:
    if mode == "complaint":
        build_complaint_dashboard_json(temp_path, source_label=original_name, write_json_file=False)
    else:
        build_dashboard_json(temp_path, source_label=original_name, write_json_file=False)

    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

    stored_path = UPLOADS_DIR / f"current_{mode}_dataset{temp_path.suffix.lower()}"
    for existing in UPLOADS_DIR.glob(f"current_{mode}_dataset.*"):
        if existing != stored_path and existing.exists():
            existing.unlink()

    if stored_path.exists():
        stored_path.unlink()
    shutil.move(str(temp_path), str(stored_path))

    save_active_dataset(stored_path, original_name, mode=mode)
    return build_dashboard_bundle(write_json_file=True)


def ensure_dashboard_payload(refresh: bool = False) -> dict[str, Any]:
    if refresh or not OUTPUT_PATH.exists():
        return build_dashboard_bundle()
    try:
        return json.loads(OUTPUT_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return build_dashboard_bundle()


def main() -> None:
    if len(sys.argv) > 1:
        dataset_path = Path(sys.argv[1]).expanduser().resolve()
        rows = load_and_validate_rows(dataset_path, mode="complaint")
        complaint_payload = build_complaint_management_payload(rows, dataset_path.name)
        general_source = resolve_active_dataset("general")
        general_payload = build_dashboard_json(general_source.path, source_label=general_source.label, write_json_file=False)
        payload = {
            "defaultMode": "general",
            "modeOptions": [
                {"key": "general", "label": "General", "available": True},
                {"key": "complaint", "label": "Complaint", "available": True},
            ],
            "modes": {"general": general_payload, "complaint": complaint_payload},
        }
        write_payload(payload)
    else:
        payload = build_dashboard_bundle()
    print(f"Wrote {OUTPUT_PATH}")
    print(f"General source: {payload['modes']['general']['meta']['sourceWorkbook']}")
    print(f"Complaint source: {payload['modes']['complaint']['meta']['sourceWorkbook']}")


if __name__ == "__main__":
    main()
